#!python 3
# blankRowInserter.py -- input from mcd the start row and the number of rows that you want to insert
# Then input the path of the xlsx.s
import openpyxl,sys,os
from pathlib import Path
#Todo:Receive the argument
startRow=int(input())
rowNumber=int(input())
xlsxPath=Path(input())

#Todo:open the xlsx file and operate
wb1=openpyxl.load_workbook(xlsxPath)
sheet1=wb1.active
wb2=openpyxl.Workbook()
sheet2=wb2.active
for i in range(1,sheet1.max_row+1):
    for j in range(1,sheet1.max_column+1):
        if i<=startRow:
            sheet2.cell(row=i,column=j).value=sheet1.cell(row=i,column=j).value
        else:
            sheet2.cell(row=i+rowNumber,column=j).value=sheet1.cell(row=i,column=j).value
wb1.close()
wb2.save(r'C:\Users\艾\Desktop\copy.xlsx')
wb2.close()
